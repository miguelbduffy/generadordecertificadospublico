from email.mime.base import MIMEBase
import smtplib, base64
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.core import exceptions
from django.http.response import JsonResponse
from django.shortcuts import redirect, render
from django.http import HttpResponse,HttpResponseRedirect
from .forms import *
from .models import *
from cryptography.fernet import Fernet
import os
import requests
import pandas as pd
from pandas import read_excel
from django.core.files.storage import FileSystemStorage
import openpyxl
import PIL
from PIL import Image, ImageDraw, ImageFont
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
import imghdr
from io import BytesIO
from email import encoders
import xlrd
from openpyxl.reader.excel import load_workbook, InvalidFileException
from openpyxl.workbook import Workbook
from openpyxl.workbook import Workbook as openpyxlWorkbook
import os
import pyexcel as p
import base64
import os
from cryptography.fernet import Fernet
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from django.conf import settings

#CHOOSE/CREATE COMPANY

@login_required(login_url='../login/')
def create_company(request, user_id):
    return render(request,"create_company.html",{"user_id":user_id})

@login_required(login_url='../login/')
def company_creation(request):
    id=request.POST.get('id')
    user=User.objects.get(id=id)
    if user != request.user:
        context = {
            logout(request)
        }
        return render(request,"not_allowed.html",{"context":context})
    else:       
        name=request.POST["name"]
        logo = request.FILES['logo']
        # fs = FileSystemStorage()
        # filename = fs.save(logo.name, logo)
        # uploaded_file_url = fs.url(filename)
        email=request.POST["email"]
        password=request.POST["password"]
        key = Fernet.generate_key()
        f = Fernet(key)
        token = f.encrypt(password.encode())

        port=request.POST["port"]
        smtp=request.POST["smtp"]

        new_company=Company()
        new_company.user=user
        new_company.name=name
        new_company.logo=logo
        new_company.email=email
        new_company.password=token
        new_company.port=port
        new_company.smtp=smtp
        new_company.key=key

        new_company.save()

        return redirect('../dashboard/'+str(user.id))

#MODIFY COMPANY
@login_required(login_url='../login/')
def modify_company(request):
    company_id = request.GET["id"]
    company = Company.objects.get(id=company_id)
    
    cut_bytes_on_key = company.key.replace("b'","").replace("'","")
    cut_bytes_on_password = company.password.replace("b'","").replace("'","")
    f = Fernet(cut_bytes_on_key.encode())
    token = (f.decrypt(cut_bytes_on_password.encode()).decode())
    company_password = token

    return render(request,"modify_company.html",{"company":company,"company_password":company_password})

@login_required(login_url='../login/')
def company_modification(request):
    id=request.POST.get('id')
    modify_company_db=Company.objects.get(id=id)
    company_owner=User.objects.get(id=modify_company_db.user.id)
    if company_owner != request.user:
        context = {
            logout(request)
        }
        return render(request,"not_allowed.html",{"context":context})
    else:
        name=request.POST.get('name')
        try:
            new_company_logo = request.FILES['new_company_logo']
            # fs = FileSystemStorage()
            # filename = fs.save(new_company_logo.name, new_company_logo)
            # uploaded_file_url = fs.url(filename)
            modify_company_db.logo=new_company_logo    
        except:
            pass
        email=request.POST.get('email')
        password=request.POST.get('password')

        port=request.POST.get('port')
        smtp=request.POST.get('smtp')
        key = Fernet.generate_key()
        fernet = Fernet(key)
        encrypted_password = fernet.encrypt(password.encode())
        modify_company_db.name=name
        modify_company_db.email=email
        modify_company_db.password=encrypted_password
        modify_company_db.port=port
        modify_company_db.smtp=smtp
        modify_company_db.key=key

        modify_company_db.save()
        return render(request,"modified_company.html",{"modify_company_db":modify_company_db})

#DELETE COMPANY
@login_required(login_url='../login/')
def delete_company(request):
    id=request.GET['id']
    company=Company.objects.get(id=id)
    company_owner=User.objects.get(id=company.user.id)
 
    if company_owner != request.user:
        context = {
            logout(request)
        }
        return render(request,"not_allowed.html",{"context":context})
    else:
        company.active=False
        company.save()
        return JsonResponse({'success':True,'id':company.id,'name':company.name})

@login_required(login_url='../login/')
def reactivate_company(request):
    id=request.GET["id"]
    company=Company.objects.get(id=id)
    company_owner=User.objects.get(id=company.user.id)
    if company_owner != request.user:
        context = {
            logout(request)
        }
        return render(request,"not_allowed.html",{"context":context})
    else:
        company.active=True
        company.save()
        return JsonResponse({'success':True,'id':company.id,'name':company.name})

#CHOOSE/CREATE COURSE
@login_required(login_url='../login/')
def choose_course(request):
    company_id=request.GET["id"]
    courses=Course.objects.filter(company=company_id,active=True)
    courses_not_active=Course.objects.filter(company=company_id,active=False)
    return render(request,"choose_course.html",{"company_id":company_id,
    "courses":courses,"courses_not_active":courses_not_active})

@login_required(login_url='../login/')
def create_course(request):
    company_id=request.GET["id"]
    form=CreateCourseForm()
    return render(request,"create_course.html",{"company_id":company_id,"form":form})

@login_required(login_url='../login/')
def course_creation(request):
    company_id=request.GET["id"]
    company_owner=Company.objects.get(id=company_id)
    name=request.GET["name"]
    description=request.GET["description"]
 
    new_course=Course()
    new_course.company=company_owner
    new_course.name=name
    new_course.description=description
    new_course.save()
    return redirect('../company/choose_course?id='+str(company_owner.id))

@login_required(login_url='../login/')
def modify_course(request):
    course_id=request.GET["id"]
    course=Course.objects.get(id=course_id)
    user=User.objects.get(id=course.company.user.id)
    user_companies=Company.objects.filter(user=user)
    if user != request.user:
        context = {
            logout(request)
        }
        return render(request,"not_allowed.html",{"context":context})
    else:
        return render(request,"modify_course.html",{"course":course,"user_companies":user_companies,"user":user})

@login_required(login_url='../login/')
def course_modification(request):
    id=request.POST.get("id")
    company=request.POST.get("company")
    company=Company.objects.get(id=company)
    name=request.POST.get("name")
    description=request.POST.get("description")
    
    course=Course.objects.get(id=id)
    course.company=company
    course.name=name
    course.description=description
    course.save()
    return render(request,"modified_course.html",{"course":course})

@login_required(login_url='../login/')
def delete_course(request):
    delete_course=request.GET["id"]
    course=Course.objects.get(id=delete_course)
    return render(request,"delete_course.html",{"delete_course":course})

@login_required(login_url='../login/')
def delete_course_confirmation(request):
    delete_course=request.GET["id"]    
    course=Course.objects.get(id=delete_course)
    if course.company.user != request.user:
        context = {
            logout(request)
        }
        return render(request,"not_allowed.html",{"context":context})
    else:
        course.active=False
        course.save()
        return JsonResponse({'success':True,"name":course.name,"id":course.id})

@login_required(login_url='../login/')
def reactivate_course(request):
    id=request.GET["id"]
    course=Course.objects.get(id=id)
    if course.company.user != request.user:
        context = {
            logout(request)
        }
        return render(request,"not_allowed.html",{"context":context})
    else:
        course.active=True
        course.save()
        return JsonResponse({"success":True,"name":course.name,"id":course.id})


#CHOOSE/CREATE SIGNEE

@login_required(login_url='../login/')
def choose_signee(request):
    course_id=request.GET["id"]
    form=CreateSigneeForm()
    course = Course.objects.get(id=course_id)
    active_company_signees=Signee.objects.filter(course__company=course.company,active=True)
    inactive_company_signees=Signee.objects.filter(course__company=course.company,active=False)

    return render(request,"choose_signee.html",{"course":course,"form":form,
    "active_company_signees":active_company_signees,
    "inactive_company_signees":inactive_company_signees})

@login_required(login_url='../login/')
def create_signee(request):
    new_signee = Signee()
    try:
        course_id=request.POST["course_id"]
        course=Course.objects.get(id=course_id)
        new_signee.course = course
    except:
        course=None

    if course.company.user != request.user:
        context = {
            logout(request)
        }
        return render(request,"not_allowed.html",{"context":context})
    else:
        if request.method == 'POST':
            form = CreateSigneeForm(
            request.POST,
            request.FILES,
            instance = new_signee,
                    )
            if form.is_valid():
                form.save()
            else:
                print(form.errors)
                return redirect('../company/choose_signee?id='+str(course.id)) 
            return JsonResponse({'success':True,'id':new_signee.id,'name':new_signee.name})
        return HttpResponse("error")

@login_required(login_url='../login/')
def modify_signee(request):
    signee_id=request.GET["id"]
    signee=Signee.objects.get(id=signee_id)
    
    return render(request,"modify_signee.html",{"signee":signee})

@login_required(login_url='../login/')
def modified_signee(request):
    signee_id=request.POST.get("signee_id")
    modify_signee_db=Signee.objects.get(id=signee_id)
    signee_owner=User.objects.get(id=modify_signee_db.course.company.user.id)
    if signee_owner != request.user:
        context = {
            logout(request)
        }
        return render(request,"not_allowed.html",{"context":context})
    else:
        try:
            new_signee_signature = request.FILES['new_signee_signature']
            # fs = FileSystemStorage()
            # filename = fs.save(new_signee_signature.name, new_signee_signature)
            # uploaded_file_url = fs.url(filename)
            modify_signee_db.signature=new_signee_signature    
        except:
            pass  
        
        signee_name=request.POST.get("signee_name")
        signee_job_title=request.POST.get("signee_job_title")
        modify_signee_db.name=signee_name
        modify_signee_db.job_title=signee_job_title

        modify_signee_db.save()
        return render(request,"modified_signee.html",{"modify_signee_db":modify_signee_db})

@login_required(login_url='../login/') 
def delete_signee(request):
    signee_id=request.GET["id"]
    signee=Signee.objects.get(id=signee_id)
    if signee.course.company.user != request.user:
        context = {
            logout(request)
        }
        return render(request,"not_allowed.html",{"context":context})
    else:
        signee.active=False
        signee.save()
        return JsonResponse({'success':True,"name":signee.name,"id":signee.id})

@login_required(login_url='../login/')
def reactivate_signee(request):
    id=request.GET["id"]
    signee=Signee.objects.get(id=id)
    if signee.course.company.user != request.user:
        context = {
            logout(request)
        }
        return render(request,"not_allowed.html",{"context":context})
    else:
        signee.active=True
        signee.save()
        return JsonResponse({'success':True,"name":signee.name,"id":signee.id})
 
#IMPORT EXCEL

@login_required(login_url='../login/') 
def upload_excel(request):
    return render(request,"import_excel.html")

def import_excel(request):
    selected_values = request.GET.getlist('signee')
    context = {}
    context["signee"] = selected_values

    if request.method == "POST":
        uploaded_file = request.FILES["excel"]
        excel_file_name=request.FILES["excel"].name

        df = pd.read_excel(uploaded_file, header=None)

        if excel_file_name[-4:] == ".xls":
            excel_file_name = excel_file_name.replace(".xls",".xlsx")
        df.to_excel(excel_file_name, index=False, header=False)
        
        fs = FileSystemStorage()
        
        wrkbk = openpyxl.load_workbook(excel_file_name)
        sh = wrkbk.active
        columns = {"Nombre":1,"Apellido":2,"DNI":3,"Mail":4,"Fecha del curso":5,"Fecha del firmante":6}
        context["url"] = fs.url(excel_file_name)
        
        for i in range(2, sh.max_row+1):
            
            print("\n")
            print("Fila ", i, " data :")
            
            name = sh.cell(row=i, column=columns["Nombre"])
            last_name = sh.cell(row=i, column=columns["Apellido"])
            id = sh.cell(row=i, column=columns["DNI"])
            convert_id = str(id.value)
            converted_id = convert_id.rstrip("0.")
            if len(converted_id) == 8:
                converted_id = converted_id[:2] + "." + converted_id[2:5] + "." + converted_id[5:8]
            elif len(converted_id) == 7:
                converted_id = converted_id[:1] + "." + converted_id[1:4] + "." + converted_id[4:7]

            mail = sh.cell(row=i, column=columns["Mail"])
            course_date = sh.cell(row=i, column=columns["Fecha del curso"])
            signature_date = sh.cell(row=i, column=columns["Fecha del firmante"])

            course_id=request.GET["course_id"]
            course=Course.objects.get(id=course_id)            
            
            get_course_signee=Signee.objects.filter(pk__in=selected_values)

            new_student = Certificate()
            new_student.course = course
            new_student.student_name = name.value
            new_student.student_last_name = last_name.value
            new_student.student_id = converted_id
            new_student.student_email = mail.value
            new_student.date_of_course = course_date.value
            new_student.signature_date = signature_date.value
            new_student.save()

            create_certificate(new_student,get_course_signee)

    return render(request,"import_excel.html", context)

#CREATE CERTIFICATE

def create_certificate(student,course_signee):
    #NOTA: EL PRIMER VALOR DE UBICACIÓN ES HORIZONTAL Y EL SEGUNDO VERTICAL.
    # DATA TO GENERATE THE CERTIFICATE
    im = Image.open(r'static/images/blank_certificate.jpg')
    d = ImageDraw.Draw(im)
    text_color = (0, 0, 0)
    # font = ImageFont.truetype("/static/arial.ttf", 25)
    # font2 = ImageFont.truetype("/static/arial.ttf", 50)
    # font3 = ImageFont.truetype("/static/arial.ttf", 40)

    #COURSE - LOCATIONS
    location_course_name = (550, 150)
    location_course_description = (550, 180)
    location_course_date_of_course = (550, 210)

    #COURSE - DATA    
    d.text(location_course_name, "Nombre del curso:" + student.course.name, fill = text_color)
    d.text(location_course_description, "Descripción del curso: " + student.course.description, fill = text_color)
    d.text(location_course_date_of_course, "Fecha del curso: " + student.date_of_course, fill = text_color)
    
    #COMPANY

    company_logo_image = Image.open(student.course.company.logo)

    im.paste(company_logo_image, (560, 30), company_logo_image)

    #STUDENT - LOCATIONS
    location_student_name = (550, 250)
    location_student_id = (550, 280)
    
    d.text(location_student_name, "Estudiante:" + student.student_name + " " + student.student_last_name, fill = text_color)
    d.text(location_student_id, "DNI: "+ student.student_id, fill = text_color)

    count = 0
    
    signature_image_2 = None

    for s in course_signee:
        new_date = student.signature_date
        if count == 0:
            location_course_signee_name = (550,600)
            location_course_signee_job_title = (550,630)
            location_course_signee_signature_date = (550,660)
            signature_image_1 = Image.open(s.signature)
        else:
            location_course_signee_name = (800,600)
            location_course_signee_job_title = (800,630)
            location_course_signee_signature_date = (800,660)
            signature_image_2 = Image.open(s.signature)

        count += 1

        d.text(location_course_signee_name, "Nombre del firmante: " + s.name, fill = text_color)
        d.text(location_course_signee_job_title, "Cargo: " + s.job_title, fill = text_color)
        d.text(location_course_signee_signature_date, "Fecha de la firma: " + new_date, fill = text_color)
        im.paste(signature_image_1, (550, 400), signature_image_1)
        if signature_image_2:
            im.paste(signature_image_2, (800, 400), signature_image_2)

    im.save("certificate_" + student.student_name + ".jpg")
    im1 = im.convert('RGB')
    im1.save(r'certificado' + '.pdf')

    send_certificate(student)

#SEND CERTIFICATE
def send_certificate(student):

    #CONFIGURATION OF EMAIL
    host = student.course.company.smtp
    port = student.course.company.port
    sender = student.course.company.email
    cut_bytes_on_key = student.course.company.key.replace("b'","").replace("'","")
    cut_bytes_on_password = student.course.company.password.replace("b'","").replace("'","")
    f = Fernet(cut_bytes_on_key.encode())
    token = (f.decrypt(cut_bytes_on_password.encode()).decode())
    password = token
    receiver = student.student_email
    
    email_conn = smtplib.SMTP(host, port)

    email_conn.ehlo()
    email_conn.starttls()
    email_conn.login(sender, password)

    the_msg = MIMEMultipart("alternative")
    the_msg['Subject'] = 'Hola ' + student.student_name + ': ' + student.course.company.name + ' te envía un certificado' 
    the_msg['From'] = str(student.course.company.name)

    the_msg['To'] = receiver

    html = """\
    <html>
    <head></head>
    <body>
        <p>Certificado adjunto.
        </p>
    </body>
    </html>
    """
    the_msg.attach(MIMEText(html,'html'))
    
    with open('certificado' + '.pdf', 'rb') as attachment:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())
    
    encoders.encode_base64(part)
    
    part.add_header(
    "Content-Disposition",
    f'attachment; filename = certificado' + '.pdf',
    )

    the_msg.attach(part)

    email_conn.sendmail(sender, receiver, the_msg.as_string())
    email_conn.close()